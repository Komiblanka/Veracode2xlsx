import xml.etree.ElementTree as ET
from openpyxl import Workbook
from sys import argv
import argparse
import os

def create_excel(Flaws):

   # Creating book and 2 sheets
    book = Workbook()
	
    stats_sheet = book.active
    stats_sheet.title = "Stats"
	
    flaws_sheet = book.create_sheet()
    flaws_sheet.title = "Flaws"
	
    libraries_sheet = book.create_sheet()
    libraries_sheet.title = "Libraries"
    
    

    # Preparing Stats sheet
    stats_sheet.cell(row = 1, column = 1).value = "Severity"
    stats_sheet.cell(row = 1, column = 2).value = "Main Client"
    stats_sheet.cell(row = 1, column = 3).value = "Third Party"
    stats_sheet.cell(row = 1, column = 4).value = "MyCompany"
    stats_sheet.cell(row = 1, column = 5).value = "?"
    stats_sheet.cell(row = 1, column = 6).value = "Fix by Policy"
    stats_sheet.cell(row = 1, column = 7).value = "Comments"

    stats_sheet.cell(row = 2, column= 1).value = 5
    stats_sheet.cell(row = 3, column= 1).value = 4
    stats_sheet.cell(row = 4, column= 1).value = 3
    stats_sheet.cell(row = 5, column= 1).value = 2
    stats_sheet.cell(row = 6, column= 1).value = 1
    stats_sheet.cell(row = 7, column= 1).value = "Total"
    
    stats_sheet.cell(row = 2, column = 2).value = '=COUNTIFS(Flaws!F:F,"5",Flaws!G:G,"Main Client")'   
    stats_sheet.cell(row = 3, column = 2).value = '=COUNTIFS(Flaws!F:F,"4",Flaws!G:G,"Main Client")'
    stats_sheet.cell(row = 4, column = 2).value = '=COUNTIFS(Flaws!F:F,"3",Flaws!G:G,"Main Client")'   
    stats_sheet.cell(row = 5, column = 2).value = '=COUNTIFS(Flaws!F:F,"2",Flaws!G:G,"Main Client")'   
    stats_sheet.cell(row = 6, column = 2).value = '=COUNTIFS(Flaws!F:F,"1",Flaws!G:G,"Main Client")'
    stats_sheet.cell(row = 7, column = 2).value = "=SUM(B2:B6)"

    stats_sheet.cell(row = 2, column = 3).value = '=COUNTIFS(Flaws!F:F,"5",Flaws!G:G,"Third Party")'
    stats_sheet.cell(row = 3, column = 3).value = '=COUNTIFS(Flaws!F:F,"4",Flaws!G:G,"Third Party")'
    stats_sheet.cell(row = 4, column = 3).value = '=COUNTIFS(Flaws!F:F,"3",Flaws!G:G,"Third Party")'
    stats_sheet.cell(row = 5, column = 3).value = '=COUNTIFS(Flaws!F:F,"2",Flaws!G:G,"Third Party")'
    stats_sheet.cell(row = 6, column = 3).value = '=COUNTIFS(Flaws!F:F,"1",Flaws!G:G,"Third Party")'
    stats_sheet.cell(row = 7, column = 3).value = "=SUM(C2:C6)"
   
    stats_sheet.cell(row = 2, column = 4).value = '=COUNTIFS(Flaws!F:F,"5",Flaws!G:G,"MyCompany")'
    stats_sheet.cell(row = 3, column = 4).value = '=COUNTIFS(Flaws!F:F,"4",Flaws!G:G,"MyCompany")'
    stats_sheet.cell(row = 4, column = 4).value = '=COUNTIFS(Flaws!F:F,"3",Flaws!G:G,"MyCompany")'
    stats_sheet.cell(row = 5, column = 4).value = '=COUNTIFS(Flaws!F:F,"2",Flaws!G:G,"MyCompany")'
    stats_sheet.cell(row = 6, column = 4).value = '=COUNTIFS(Flaws!F:F,"1",Flaws!G:G,"MyCompany")'
    stats_sheet.cell(row = 7, column = 4).value = "=SUM(D2:D6)"
    
    stats_sheet.cell(row = 2, column = 5).value = '=COUNTIFS(Flaws!F:F,"5",Flaws!G:G,"?")'
    stats_sheet.cell(row = 3, column = 5).value = '=COUNTIFS(Flaws!F:F,"4",Flaws!G:G,"?")'
    stats_sheet.cell(row = 4, column = 5).value = '=COUNTIFS(Flaws!F:F,"3",Flaws!G:G,"?")'
    stats_sheet.cell(row = 5, column = 5).value = '=COUNTIFS(Flaws!F:F,"2",Flaws!G:G,"?")'
    stats_sheet.cell(row = 6, column = 5).value = '=COUNTIFS(Flaws!F:F,"1",Flaws!G:G,"?")'
    stats_sheet.cell(row = 7, column = 5).value = "=SUM(E2:E6)"
    
    stats_sheet.cell(row = 2, column = 6).value = 0
    stats_sheet.cell(row = 3, column = 6).value = 0
    stats_sheet.cell(row = 4, column = 6).value = 0
    stats_sheet.cell(row = 5, column = 6).value = 0
    stats_sheet.cell(row = 6, column = 6).value = 0
    stats_sheet.cell(row = 7, column = 6).value = "=SUM(F2:F6)"     
      
    # Preparing Flaws sheet
    
    flaws_sheet.cell(row = 1, column = 1).value = "Flaw id"
    flaws_sheet.cell(row = 1, column = 2).value = "Category"
    flaws_sheet.cell(row = 1, column = 3).value = "Sub Category"
    flaws_sheet.cell(row = 1, column = 4).value = "Library"
    flaws_sheet.cell(row = 1, column = 5).value = "File Path"
    flaws_sheet.cell(row = 1, column = 6).value = "Line"
    flaws_sheet.cell(row = 1, column = 7).value = "Severity"
    flaws_sheet.cell(row = 1, column = 8).value = "Propietary"
    

    # Preparing Libraries sheet
    
    libraries_sheet.cell(row = 1, column = 1).value = "Library"
    libraries_sheet.cell(row = 1, column = 2).value = "Propietary"
    
    
    
    vulnerable_libraries = set()
    counter = 1
    
    # Writing all the flaws
    for flaw_category in Flaws:
        counter += 1
        flaw = flaw_category[0]
        category = flaw_category[1]
        compiled = False
        
        # Flaw id
        flaws_sheet.cell(row = counter, column = 1).value = int(flaw.get("issueid"))
        # Category
        flaws_sheet.cell(row = counter, column = 2).value = category
        # Sub Category
        flaws_sheet.cell(row = counter, column = 3).value = flaw.get("categoryname")
        # Library
        flaws_sheet.cell(row = counter, column = 4).value = flaw.get("module")
        # File path
        pathfile = flaw.get("sourcefilepath") + flaw.get("sourcefile")
        
        if not pathfile:
            compiled = True
        if compiled:
            pathfile = "Compiled library (check report): " + flaw.get("module")
        flaws_sheet.cell(row = counter, column = 5).value = pathfile
        vulnerable_libraries.add(flaw.get("module"))
        # Line
        if not compiled:
            flaws_sheet.cell(row = counter, column = 6).value = int(flaw.get("line"))
        else:
            flaws_sheet.cell(row = counter, column = 6).value = "Vulnerable class: " + flaw.get("functionprototype") + " at " + flaw.get("functionrelativelocation") + "%"
        # Severity
        flaws_sheet.cell(row = counter, column = 7).value = int(flaw.get("severity"))
        # Propietary
        flaws_sheet.cell(row = counter, column = 8).value = "=VLOOKUP(D" + str(counter) + ",Libraries!A:B,2,FALSE)"
    
    
    counter = 1
    
    vulnerable_libraries = list(vulnerable_libraries)
    vulnerable_libraries.sort()
    
    libraries = set()
 
    # Writing all vulnerable libraries
    for library in vulnerable_libraries:
        counter += 1
        libraries_sheet.cell(row = counter, column = 1).value = library # File
    
    
    
    return book




parser = argparse.ArgumentParser(description='Transforms a Veracode report in XML format to an .xlsx file.')

parser.add_argument('infile', help='XML report you want to parse')
parser.add_argument('outfile', help='name of the output excel file in xlsx format')

args = parser.parse_args()

inxml = args.infile
outxls = args.outfile

if not(os.path.isfile(inxml)):
    print inxml + "doesn't exist."
    exit()

# Preparing to parse XML
tree = ET.parse(inxml)
root = tree.getroot()

Flaws = []

print ""
print "Parsing..."
# Go through the XML and create a list of lists of flaws to be able to associate a flaw to a category. Please note the namespace when finding nodes
for severity in root.findall("{https://www.veracode.com/schema/reports/export/1.0}severity"):
    for category in severity:
        for flaw in category.iter("{https://www.veracode.com/schema/reports/export/1.0}flaw"):
            Flaws.append([flaw, category.get("categoryname")])

print "Parsing done correctly."

print "Creating excel file..."
print ""

book = create_excel(Flaws)

book.save(outxls)

print "Excel file: " + outxls + " built correctly!"
print ""
print "Now fill in the propietary column for each of the libraries!"
print ""

